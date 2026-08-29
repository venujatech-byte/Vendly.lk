from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.core.errors import ApiError
from app.services.operations_service import (
    build_orders_workbook,
    validate_report_payload,
)


def test_report_payload_rejects_unknown_type():
    with pytest.raises(ApiError) as error:
        validate_report_payload(
            {"type": "unknown"},
            {"delayed", "lost"},
            "Courier issue type",
        )

    assert error.value.status_code == 422


def test_order_export_creates_real_excel_workbook():
    stream = build_orders_workbook(
        [
            {
                "orderNumber": "VD-000001",
                "customerSnapshot": {
                    "name": "Kamal",
                    "normalizedPhone": "94771234567",
                },
                "deliveryAddress": {"line1": "10 Main Road", "district": "Colombo"},
                "items": [{"name": "Watch", "size": "", "quantity": 1}],
                "itemCount": 1,
                "subtotalMinor": 200000,
                "deliveryFeeMinor": 45000,
                "totalAmountMinor": 245000,
                "courierSnapshot": {"name": "Courier One"},
                "fulfilmentStatus": "confirmed",
            },
        ],
    )
    workbook = load_workbook(stream)
    sheet = workbook["Orders"]

    # The courier template keeps Waybill Id in column A and Order Number in B.
    assert sheet["A1"].value == "Waybill Id"
    assert sheet["B1"].value == "Order Number"
    assert sheet["B2"].value == "VD-000001"
    assert sheet["C2"].value == "Kamal"
    assert sheet["H2"].value == 2450
    assert sheet["J2"].value == 2000


def test_order_export_fills_the_courier_template_without_replacing_its_layout():
    template = Workbook()
    sheet = template.active
    sheet.title = "Courier Upload"
    sheet["A1"] = "Courier upload instructions"
    sheet.append(["Waybill No", "Receiver Name", "Order Number", "COD"])
    sheet.append(["sample", "sample", "sample", 0])
    template_stream = BytesIO()
    template.save(template_stream)

    stream = build_orders_workbook(
        [
            {
                "waybillNumber": "KMB-00012",
                "orderNumber": "VD-000012",
                "customerSnapshot": {"name": "Nimali"},
                "totalAmountMinor": 345000,
            },
        ],
        template_bytes=template_stream.getvalue(),
    )
    exported = load_workbook(stream)
    exported_sheet = exported["Courier Upload"]

    assert exported_sheet["A1"].value == "Courier upload instructions"
    assert exported_sheet["A3"].value == "KMB-00012"
    assert exported_sheet["B3"].value == "Nimali"
    assert exported_sheet["C3"].value == "VD-000012"
    assert exported_sheet["D3"].value == 3450


def test_order_export_can_be_limited_to_selected_order_ids(monkeypatch):
    from app.services import operations_service

    orders = [
        {"id": "order-1", "orderNumber": "VD-000001"},
        {"id": "order-2", "orderNumber": "VD-000002"},
    ]
    captured = {}

    monkeypatch.setattr(
        operations_service,
        "get_courier_export_template",
        lambda *_args: {"content": b"courier-template"},
    )
    monkeypatch.setattr(
        operations_service,
        "list_orders",
        lambda *_args, **_kwargs: orders,
    )

    def fake_build(exported_orders, template_bytes=None):
        captured["orders"] = exported_orders
        captured["template"] = template_bytes
        return BytesIO()

    monkeypatch.setattr(operations_service, "build_orders_workbook", fake_build)

    operations_service.export_orders(
        object(),
        "business-1",
        courier_id="courier-1",
        order_ids=["order-2"],
    )

    assert captured["orders"] == [orders[1]]
    assert captured["template"] == b"courier-template"


class FakeDocument:
    """Mirrors the .document(id).collection(name) chain the service walks."""

    def __init__(self, store, key):
        self.store = store
        self.key = key

    def collection(self, _name):
        return FakeCollection(self.store)

    def set(self, data):
        self.store[self.key] = data

    def delete(self):
        self.store.pop(self.key, None)


class FakeCollection:
    def __init__(self, store):
        self.store = store

    def document(self, name):
        return FakeDocument(self.store, name)


class FakeDatabase(FakeCollection):
    def collection(self, _name):
        return FakeCollection(self.store)


def ai_failure(kind="configuration", at="2026-08-27T01:52:21Z"):
    return {
        "failure": {"kind": kind, "provider": "groq", "model": "a-model", "at": at},
    }


def test_an_ai_failure_raises_one_notification(monkeypatch):
    from app.services import operations_service

    monkeypatch.setattr(operations_service, "_SYNCED_AI_FAILURES", {})
    store = {}
    operations_service.sync_ai_failure_notification(
        FakeDatabase(store),
        "biz",
        ai_failure(),
    )

    assert operations_service.AI_STATUS_NOTIFICATION_ID in store
    notification = store[operations_service.AI_STATUS_NOTIFICATION_ID]
    assert notification["type"] == "ai-status"
    assert notification["isRead"] is False
    assert "a-model" in notification["message"]


def test_a_repeated_failure_does_not_write_again(monkeypatch):
    from app.services import operations_service

    monkeypatch.setattr(operations_service, "_SYNCED_AI_FAILURES", {})
    store = {}
    database = FakeDatabase(store)

    operations_service.sync_ai_failure_notification(database, "biz", ai_failure())
    store.clear()
    # A dead provider fails on every single message; it must not write each time.
    operations_service.sync_ai_failure_notification(database, "biz", ai_failure())

    assert store == {}


def test_recovery_clears_the_notification(monkeypatch):
    from app.services import operations_service

    monkeypatch.setattr(operations_service, "_SYNCED_AI_FAILURES", {})
    store = {}
    database = FakeDatabase(store)

    operations_service.sync_ai_failure_notification(database, "biz", ai_failure())
    operations_service.sync_ai_failure_notification(database, "biz", {"failure": None})

    # A stale warning is how a notification becomes something people ignore.
    assert store == {}


def test_a_new_failure_after_recovery_notifies_again(monkeypatch):
    from app.services import operations_service

    monkeypatch.setattr(operations_service, "_SYNCED_AI_FAILURES", {})
    store = {}
    database = FakeDatabase(store)

    operations_service.sync_ai_failure_notification(database, "biz", ai_failure())
    operations_service.sync_ai_failure_notification(database, "biz", {"failure": None})
    operations_service.sync_ai_failure_notification(
        database,
        "biz",
        ai_failure(at="2026-08-27T09:00:00Z"),
    )

    assert operations_service.AI_STATUS_NOTIFICATION_ID in store
