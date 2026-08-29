"""Excel export/import helpers for analytics and inventory.

The inventory workbook is deliberately split into Categories, Products and
Variants sheets.  This keeps it readable in Excel while preserving nested
product media and size/variant rows for a safe round trip back into Vendly.
"""

from io import BytesIO
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.core.errors import ApiError
from app.core.serialization import serialize_snapshot
from app.services.category_service import create_category
from app.services.product_service import create_product, update_product


MONEY_FORMAT = '#,##0.00'
HEADER_FILL = PatternFill("solid", fgColor="0B3B6E")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _amount(minor_units):
    return round(int(minor_units or 0) / 100, 2)


def _as_bool(value):
    if isinstance(value, bool):
        return value
    return str(value or "").strip().casefold() in {"1", "true", "yes", "y"}


def _as_number(value, default=0):
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError) as error:
        raise ApiError("invalid_inventory_workbook", f"Invalid number: {value}", 422) from error


def _as_integer(value, default=0):
    number = _as_number(value, default)
    if not float(number).is_integer():
        raise ApiError("invalid_inventory_workbook", f"Expected a whole number: {value}", 422)
    return int(number)


def _style_sheet(sheet, widths):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _workbook_stream(workbook):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def export_ledger_workbook(entries):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transaction Ledger"
    sheet.append([
        "Date and time", "Reference", "Transaction type", "Customer",
        "Details", "Payment method", "Payment status", "Status",
        "Money in (LKR)", "Money out (LKR)", "Running balance (LKR)",
    ])
    for entry in entries:
        is_credit = entry.get("direction") == "credit"
        sheet.append([
            entry.get("createdAt", ""), entry.get("reference", ""),
            entry.get("label", ""), entry.get("customerName", ""),
            entry.get("description", ""), entry.get("paymentMethod", ""),
            entry.get("paymentStatus", ""), entry.get("status", ""),
            _amount(entry.get("amountMinor")) if is_credit else None,
            _amount(entry.get("amountMinor")) if not is_credit else None,
            _amount(entry.get("balanceMinor")),
        ])
    _style_sheet(sheet, {
        "A": 23, "B": 18, "C": 22, "D": 24, "E": 38, "F": 18,
        "G": 18, "H": 18, "I": 18, "J": 18, "K": 22,
    })
    for row in sheet.iter_rows(min_row=2, min_col=9, max_col=11):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    return _workbook_stream(workbook)


def export_inventory_workbook(database, business_id, product_ids=None):
    business = database.collection("businesses").document(business_id)
    categories = [serialize_snapshot(row) for row in business.collection("categories").stream()]
    products = [serialize_snapshot(row) for row in business.collection("products").stream()]
    selected_ids = {str(product_id) for product_id in (product_ids or []) if product_id}
    if selected_ids:
        products = [product for product in products if product.get("id") in selected_ids]

    workbook = Workbook()
    products_sheet = workbook.active
    products_sheet.title = "Products"
    products_sheet.append([
        "Product ID", "Product Name", "Colour Name", "Colour Hex",
        "Product Type", "Product Size", "Category Name", "Brand",
        "Supplier ID", "Description", "AI Description", "Tax Category",
        "Has Variants", "SKU Prefix", "Cost Price (LKR)",
        "Selling Price (LKR)", "Compare At Price (LKR)", "Weight (kg)",
        "Low Stock Threshold", "Warranty Months", "Warranty Notes", "Status",
        "Media JSON", "Image URLs", "Video URLs",
    ])
    for product in products:
        media = product.get("media") or []
        images = [item.get("url", "") for item in media if item.get("type", "image") == "image" and item.get("url")]
        videos = [item.get("url", "") for item in media if item.get("type") == "video" and item.get("url")]
        products_sheet.append([
            product.get("id", ""), product.get("name", ""),
            product.get("colourName", ""), product.get("colourHex", ""),
            product.get("productType", ""), product.get("productSize", ""),
            product.get("categoryName", "Uncategorized"), product.get("brand", ""),
            product.get("supplierId", ""), product.get("description", ""),
            product.get("aiDescription", ""), product.get("taxCategory", "standard"),
            bool(product.get("hasSizes")), product.get("skuPrefix", ""),
            _amount(product.get("costPriceMinor")), _amount(product.get("sellingPriceMinor")),
            _amount(product.get("compareAtPriceMinor")), round(int(product.get("weightGrams") or 0) / 1000, 3),
            product.get("lowStockThreshold", 5), product.get("warrantyPeriodMonths", 0),
            product.get("warrantyNotes", ""), product.get("status", "active"),
            json.dumps(media, ensure_ascii=False), "\n".join(images), "\n".join(videos),
        ])

    variants_sheet = workbook.create_sheet("Variants")
    variants_sheet.append([
        "Product ID", "Product Name", "Variant ID", "Variant / Size", "SKU",
        "Barcode", "Stock On Hand", "Reserved Stock", "Available Stock",
        "Cost Price (LKR)", "Selling Price (LKR)", "Weight (kg)", "Image URL", "Status",
    ])
    for product in products:
        for variant in product.get("variantSummaries") or []:
            variants_sheet.append([
                product.get("id", ""), product.get("name", ""), variant.get("id", ""),
                variant.get("size", ""), variant.get("sku", ""), variant.get("barcode", ""),
                variant.get("stockOnHand", variant.get("stockAvailable", 0)),
                variant.get("stockReserved", 0), variant.get("stockAvailable", 0),
                _amount(variant.get("costPriceMinor", product.get("costPriceMinor"))),
                _amount(variant.get("sellingPriceMinor", product.get("sellingPriceMinor"))),
                round(int(variant.get("weightGrams", product.get("weightGrams")) or 0) / 1000, 3),
                variant.get("imageUrl", ""), variant.get("status", "active"),
            ])

    categories_sheet = workbook.create_sheet("Categories")
    categories_sheet.append(["Category ID", "Category Name", "Description", "Status", "Sort Order"])
    for category in categories:
        categories_sheet.append([
            category.get("id", ""), category.get("name", ""), category.get("description", ""),
            category.get("status", "active"), category.get("sortOrder", 0),
        ])

    _style_sheet(products_sheet, {"A": 22, "B": 28, "G": 22, "J": 42, "K": 42, "W": 45, "X": 45, "Y": 45})
    _style_sheet(variants_sheet, {"A": 22, "B": 28, "C": 22, "D": 18, "E": 20, "F": 20, "M": 45})
    _style_sheet(categories_sheet, {"A": 22, "B": 26, "C": 45, "D": 15, "E": 14})
    for sheet in (products_sheet, variants_sheet):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if "Price (LKR)" in str(sheet.cell(1, cell.column).value):
                    cell.number_format = MONEY_FORMAT
    return _workbook_stream(workbook)


def _records(sheet):
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    return [
        dict(zip(headers, row))
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in row)
    ]


def parse_inventory_workbook(stream):
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except Exception as error:
        raise ApiError("invalid_inventory_workbook", "Choose a valid Vendly .xlsx inventory file.", 422) from error
    required_sheets = {"Products", "Variants", "Categories"}
    if not required_sheets.issubset(workbook.sheetnames):
        raise ApiError("invalid_inventory_workbook", "The workbook must contain Products, Variants and Categories sheets.", 422)

    product_rows = _records(workbook["Products"])
    variant_rows = _records(workbook["Variants"])
    category_rows = _records(workbook["Categories"])
    variants_by_product = {}
    seen_skus, seen_barcodes = set(), set()
    for row in variant_rows:
        product_id = str(row.get("Product ID") or "").strip()
        sku = str(row.get("SKU") or "").strip().upper()
        barcode = str(row.get("Barcode") or "").strip()
        if not product_id or not sku or not barcode:
            raise ApiError("invalid_inventory_workbook", "Every variant row needs Product ID, SKU and Barcode.", 422)
        if sku in seen_skus or barcode in seen_barcodes:
            raise ApiError("invalid_inventory_workbook", f"Duplicate SKU or barcode in workbook: {sku or barcode}", 422)
        seen_skus.add(sku)
        seen_barcodes.add(barcode)
        variants_by_product.setdefault(product_id, []).append({
            "id": str(row.get("Variant ID") or "").strip(),
            "size": str(row.get("Variant / Size") or "").strip(),
            "sku": sku,
            "barcode": barcode,
            "stock": _as_integer(row.get("Stock On Hand")),
            "costPrice": _as_number(row.get("Cost Price (LKR)")),
            "sellingPrice": _as_number(row.get("Selling Price (LKR)")),
            "imageUrl": str(row.get("Image URL") or "").strip(),
        })

    products = []
    for row in product_rows:
        product_id = str(row.get("Product ID") or "").strip()
        variants = variants_by_product.get(product_id, [])
        if not product_id or not variants:
            raise ApiError("invalid_inventory_workbook", f"Product {row.get('Product Name') or product_id or '?'} has no variant rows.", 422)
        media_text = str(row.get("Media JSON") or "").strip()
        try:
            media = json.loads(media_text) if media_text else []
        except json.JSONDecodeError as error:
            raise ApiError("invalid_inventory_workbook", f"Invalid Media JSON for {row.get('Product Name')}.", 422) from error
        products.append({
            "sourceProductId": product_id,
            "categoryName": str(row.get("Category Name") or "Uncategorized").strip(),
            "status": str(row.get("Status") or "active").strip().lower(),
            "payload": {
                "name": str(row.get("Product Name") or "").strip(),
                "colourName": str(row.get("Colour Name") or "").strip(),
                "colourHex": str(row.get("Colour Hex") or "").strip(),
                "productType": str(row.get("Product Type") or "").strip(),
                "productSize": str(row.get("Product Size") or "").strip(),
                "brand": str(row.get("Brand") or "").strip(),
                "supplierId": str(row.get("Supplier ID") or "").strip(),
                "description": str(row.get("Description") or "").strip(),
                "aiDescription": str(row.get("AI Description") or "").strip(),
                "taxCategory": str(row.get("Tax Category") or "standard").strip(),
                "hasSizes": _as_bool(row.get("Has Variants")),
                "skuPrefix": str(row.get("SKU Prefix") or "").strip(),
                "costPrice": _as_number(row.get("Cost Price (LKR)")),
                "sellingPrice": _as_number(row.get("Selling Price (LKR)")),
                "compareAtPrice": _as_number(row.get("Compare At Price (LKR)")),
                "weightKg": _as_number(row.get("Weight (kg)")),
                "lowStockThreshold": _as_integer(row.get("Low Stock Threshold"), 5),
                "warrantyPeriodMonths": _as_integer(row.get("Warranty Months")),
                "warrantyNotes": str(row.get("Warranty Notes") or "").strip(),
                "media": media,
                "variants": variants,
            },
        })
    return {"categories": category_rows, "products": products}


def import_inventory_workbook(database, business_id, uid, stream):
    imported = parse_inventory_workbook(stream)
    business = database.collection("businesses").document(business_id)
    existing_categories = [serialize_snapshot(row) for row in business.collection("categories").stream()]
    category_by_name = {str(row.get("name", "")).casefold(): row for row in existing_categories}
    categories_created = 0
    for row in imported["categories"]:
        name = str(row.get("Category Name") or "").strip()
        if not name or name.casefold() in category_by_name:
            continue
        created = create_category(database, business_id, {
            "name": name,
            "description": str(row.get("Description") or "").strip(),
            "sortOrder": _as_integer(row.get("Sort Order")),
        })
        category_by_name[name.casefold()] = created
        categories_created += 1

    existing_products = [serialize_snapshot(row) for row in business.collection("products").stream()]
    product_by_id = {row["id"]: row for row in existing_products}
    product_by_sku = {
        str(variant.get("sku", "")).upper(): product
        for product in existing_products
        for variant in product.get("variantSummaries") or []
        if variant.get("sku")
    }
    created_count = updated_count = 0
    for imported_product in imported["products"]:
        payload = imported_product["payload"]
        category = category_by_name.get(imported_product["categoryName"].casefold())
        payload["categoryId"] = category.get("id", "") if category else ""
        payload["status"] = imported_product["status"] if imported_product["status"] in {"active", "draft", "archived"} else "active"
        first_sku = payload["variants"][0]["sku"]
        existing = product_by_id.get(imported_product["sourceProductId"]) or product_by_sku.get(first_sku)
        if existing:
            update_product(database, business_id, existing["id"], payload)
            updated_count += 1
        else:
            created = create_product(database, business_id, uid, payload)
            if payload["status"] != "active":
                update_product(database, business_id, created["id"], {"status": payload["status"]})
            created_count += 1
    return {
        "productsCreated": created_count,
        "productsUpdated": updated_count,
        "categoriesCreated": categories_created,
        "productCount": len(imported["products"]),
    }
