from io import BytesIO

from firebase_admin import firestore
from google.cloud import firestore as google_firestore
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.core.errors import ApiError
from app.core.serialization import serialize_snapshot
from app.services.order_service import get_order, list_orders
from app.services.courier_service import get_courier_export_template
from app.services.fraud_service import (
    global_fraud_reference,
    global_registry_increment,
)
from app.services.text import optional_text, required_text


FRAUD_REASONS = {
    "fake-details",
    "invalid-address",
    "no-contact",
    "refused-order",
    "repeat-return",
    "other",
}
COURIER_ISSUE_TYPES = {
    "branch-problem",
    "delayed",
    "damaged",
    "lost",
    "other",
}
WAYBILL_ALLOWED_STATUSES = {"confirmed", "packed", "shipped", "delivered"}

ORDER_EXPORT_COLUMNS = {
    "waybill": ("waybill id", "waybill no", "waybill number"),
    "order": ("order number", "order id", "order no"),
    "receiver": ("receiver name", "customer name", "customer"),
    "address": ("delivery address", "address"),
    "district": ("district name", "district"),
    "city": ("city", "nearest city"),
    "phone": ("receiver phone", "phone", "phone number"),
    "cod": ("cod", "cash on delivery", "amount to collect"),
    "description": ("description", "items", "item names"),
    "actual": ("actual value", "subtotal", "item value"),
}


def validate_report_payload(payload, allowed_values, field_label):
    try:
        report_type = required_text(payload.get("type"), field_label, 60).lower()
        note = optional_text(payload.get("note"), 1000)
    except ValueError as error:
        raise ApiError("validation_error", str(error), 422) from error

    if report_type not in allowed_values:
        raise ApiError(
            "validation_error",
            f"Choose a valid {field_label.lower()}.",
            422,
            {"allowedValues": sorted(allowed_values)},
        )

    return report_type, note


def create_fraud_report(database, business_id, order_id, uid, payload):
    reason, note = validate_report_payload(payload, FRAUD_REASONS, "Fraud reason")
    business_reference = database.collection("businesses").document(business_id)
    order_reference = business_reference.collection("orders").document(order_id)
    report_reference = business_reference.collection("fraudReports").document(order_id)
    notification_reference = business_reference.collection("notifications").document()
    transaction = database.transaction()

    @google_firestore.transactional
    def create_in_transaction(current_transaction):
        order_snapshot = order_reference.get(transaction=current_transaction)
        report_snapshot = report_reference.get(transaction=current_transaction)

        if not order_snapshot.exists:
            raise ApiError("order_not_found", "Order not found.", 404)
        if report_snapshot.exists:
            raise ApiError(
                "fraud_report_exists",
                "This order has already been reported.",
                409,
            )

        order = order_snapshot.to_dict()
        customer_reference = business_reference.collection("customers").document(
            order.get("customerId", ""),
        )
        customer_snapshot = customer_reference.get(transaction=current_transaction)
        customer_phone = order.get("customerSnapshot", {}).get("normalizedPhone", "")
        registry_reference = (
            global_fraud_reference(database, customer_phone)
            if customer_phone
            else None
        )
        timestamp = firestore.SERVER_TIMESTAMP
        report = {
            "orderId": order_id,
            "orderNumber": order.get("orderNumber", ""),
            "customerId": order.get("customerId", ""),
            "customerSnapshot": order.get("customerSnapshot", {}),
            "reason": reason,
            "note": note,
            "status": "active",
            "reportedBy": uid,
            "createdAt": timestamp,
            "updatedAt": timestamp,
        }
        current_transaction.set(report_reference, report)
        if registry_reference:
            current_transaction.set(
                registry_reference,
                global_registry_increment(report_count=1),
                merge=True,
            )
        current_transaction.update(
            order_reference,
            {
                "fraudReport": {
                    "reason": reason,
                    "status": "active",
                    "reportedBy": uid,
                },
                "updatedAt": timestamp,
            },
        )

        if customer_snapshot.exists:
            customer = customer_snapshot.to_dict()
            current_transaction.update(
                customer_reference,
                {
                    "fraudReportCount": customer.get("fraudReportCount", 0) + 1,
                    "riskLevel": "high",
                    "tags": firestore.ArrayUnion(["fraud-reported"]),
                    "updatedAt": timestamp,
                },
            )

        current_transaction.set(
            notification_reference,
            {
                "type": "fraud-report",
                "title": f"Fraud report for {order.get('orderNumber', 'order')}",
                "message": f"Reason: {reason.replace('-', ' ')}.",
                "orderId": order_id,
                "orderNumber": order.get("orderNumber", ""),
                "isRead": False,
                "createdAt": timestamp,
            },
        )

    create_in_transaction(transaction)
    return serialize_snapshot(report_reference.get())


def record_courier_issue(database, business_id, order_id, uid, payload):
    issue_type, note = validate_report_payload(
        payload,
        COURIER_ISSUE_TYPES,
        "Courier issue type",
    )
    business_reference = database.collection("businesses").document(business_id)
    order_reference = business_reference.collection("orders").document(order_id)
    issue_reference = business_reference.collection("courierIssues").document()
    transaction = database.transaction()

    @google_firestore.transactional
    def create_in_transaction(current_transaction):
        order_snapshot = order_reference.get(transaction=current_transaction)

        if not order_snapshot.exists:
            raise ApiError("order_not_found", "Order not found.", 404)

        order = order_snapshot.to_dict()
        courier_id = order.get("courierId")

        if not courier_id:
            raise ApiError(
                "courier_not_assigned",
                "Assign a courier before reporting a courier issue.",
                409,
            )

        courier_reference = business_reference.collection("couriers").document(
            courier_id,
        )
        courier_snapshot = courier_reference.get(transaction=current_transaction)

        if not courier_snapshot.exists:
            raise ApiError("courier_not_found", "Courier not found.", 404)

        courier = courier_snapshot.to_dict()
        district = order.get("district", "unknown")
        district_counts = dict(courier.get("districtIssueCounts", {}))
        district_counts[district] = district_counts.get(district, 0) + 1
        timestamp = firestore.SERVER_TIMESTAMP
        current_transaction.set(
            issue_reference,
            {
                "orderId": order_id,
                "orderNumber": order.get("orderNumber", ""),
                "courierId": courier_id,
                "courierName": courier.get("name", ""),
                "district": district,
                "type": issue_type,
                "note": note,
                "reportedBy": uid,
                "createdAt": timestamp,
            },
        )
        current_transaction.update(
            courier_reference,
            {
                "districtIssueCounts": district_counts,
                "issueCount": courier.get("issueCount", 0) + 1,
                "updatedAt": timestamp,
            },
        )
        current_transaction.update(
            order_reference,
            {
                "courierIssueCount": order.get("courierIssueCount", 0) + 1,
                "updatedAt": timestamp,
            },
        )

    create_in_transaction(transaction)
    return serialize_snapshot(issue_reference.get())


def generate_waybill(database, business_id, order_id, uid):
    business_reference = database.collection("businesses").document(business_id)
    order_reference = business_reference.collection("orders").document(order_id)
    waybill_reference = business_reference.collection("waybills").document(order_id)
    transaction = database.transaction()

    @google_firestore.transactional
    def generate_in_transaction(current_transaction):
        business_snapshot = business_reference.get(transaction=current_transaction)
        order_snapshot = order_reference.get(transaction=current_transaction)

        if not order_snapshot.exists:
            raise ApiError("order_not_found", "Order not found.", 404)

        order = order_snapshot.to_dict()

        if order.get("waybillNumber"):
            return
        if order.get("fulfilmentStatus") not in WAYBILL_ALLOWED_STATUSES:
            raise ApiError(
                "order_not_ready_for_waybill",
                "Confirm the order before generating a waybill.",
                409,
            )

        business = business_snapshot.to_dict() if business_snapshot.exists else {}
        courier_reference = business_reference.collection("couriers").document(order.get("courierId", ""))
        courier_snapshot = courier_reference.get(transaction=current_transaction) if order.get("courierId") else None
        courier = courier_snapshot.to_dict() if courier_snapshot and courier_snapshot.exists else {}
        sequence = courier.get("nextWaybillSequence", courier.get("waybillStart", business.get("nextWaybillSequence", 1)))
        waybill_end = courier.get("waybillEnd", 999999)
        if sequence > waybill_end:
            raise ApiError("waybill_range_exhausted", "This courier's waybill range is exhausted.", 409)
        waybill_number = f"{courier.get('waybillPrefix', 'VWB')}-{sequence:08d}"
        timestamp = firestore.SERVER_TIMESTAMP
        current_transaction.set(
            waybill_reference,
            {
                "waybillNumber": waybill_number,
                "orderId": order_id,
                "orderNumber": order.get("orderNumber", ""),
                "courierId": order.get("courierId", ""),
                "courierSnapshot": order.get("courierSnapshot", {}),
                "customerSnapshot": order.get("customerSnapshot", {}),
                "deliveryAddress": order.get("deliveryAddress", {}),
                "totalWeightGrams": order.get("totalWeightGrams", 0),
                "generatedBy": uid,
                "createdAt": timestamp,
            },
        )
        current_transaction.update(
            order_reference,
            {"waybillNumber": waybill_number, "updatedAt": timestamp},
        )
        current_transaction.update(
            business_reference,
            {"nextWaybillSequence": sequence + 1, "updatedAt": timestamp},
        )
        if courier_snapshot and courier_snapshot.exists:
            current_transaction.update(
                courier_reference,
                {"nextWaybillSequence": sequence + 1, "updatedAt": timestamp},
            )

    generate_in_transaction(transaction)
    return get_order(database, business_id, order_id)


def format_address(address):
    return ", ".join(
        str(address.get(field, "")).strip()
        for field in ("line1", "line2", "city", "district", "postalCode", "country")
        if str(address.get(field, "")).strip()
    )


def order_export_values(order):
    items = order.get("items", [])
    address = order.get("deliveryAddress", {}) or {}
    customer = order.get("customerSnapshot", {}) or {}
    phone_numbers = [
        customer.get("phoneNumber") or customer.get("normalizedPhone") or "",
        customer.get("secondaryPhoneNumber")
        or customer.get("normalizedSecondaryPhone")
        or "",
    ]
    total_minor = order.get("totalAmountMinor", 0) or 0
    delivery_minor = order.get("deliveryFeeMinor", 0) or 0
    paid_minor = order.get("paidAmountMinor", 0) or 0
    return {
        "waybill": order.get("waybillNumber", ""),
        "order": order.get("orderNumber") or order.get("id", ""),
        "receiver": customer.get("name", ""),
        "address": format_address(address),
        "district": address.get("district", ""),
        "city": address.get("city", ""),
        "phone": ", ".join(phone for phone in phone_numbers if phone),
        "cod": (total_minor - paid_minor) / 100,
        "description": ", ".join(
            f"{item.get('name', '')} {item.get('size', '')}".strip()
            for item in items
            if item.get("name")
        ),
        "actual": (total_minor - delivery_minor) / 100,
    }


def normalize_heading(value):
    return " ".join(str(value or "").strip().casefold().replace("_", " ").split())


def find_template_columns(workbook):
    aliases = {
        alias: field
        for field, field_aliases in ORDER_EXPORT_COLUMNS.items()
        for alias in field_aliases
    }
    for sheet in workbook.worksheets:
        for row_number in range(1, min(sheet.max_row, 30) + 1):
            columns = {}
            for cell in sheet[row_number]:
                field = aliases.get(normalize_heading(cell.value))
                if field:
                    columns[field] = cell.column
            if "order" in columns and len(columns) >= 3:
                return sheet, row_number, columns
    raise ApiError(
        "invalid_export_template",
        "The courier template needs recognizable headings such as Order Number, Receiver Name and COD.",
        422,
    )


def build_orders_from_template(orders, template_bytes):
    workbook = load_workbook(BytesIO(template_bytes))
    sheet, header_row, columns = find_template_columns(workbook)
    first_data_row = header_row + 1

    for row_number in range(first_data_row, sheet.max_row + 1):
        for column_number in columns.values():
            sheet.cell(row=row_number, column=column_number).value = None

    for index, order in enumerate(orders):
        values = order_export_values(order)
        for field, column_number in columns.items():
            sheet.cell(
                row=first_data_row + index,
                column=column_number,
            ).value = values[field]

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_orders_workbook(orders, template_bytes=None):
    if template_bytes:
        return build_orders_from_template(orders, template_bytes)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    headers = [
        "Waybill Id",
        "Order Number",
        "Receiver Name",
        "Delivery Address",
        "District Name",
        "City",
        "Receiver Phone",
        "COD",
        "Description",
        "Actual Value",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="0B3B6E")

    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for order in orders:
        items = order.get("items", [])
        address = order.get("deliveryAddress", {}) or {}
        customer = order.get("customerSnapshot", {}) or {}
        phone_numbers = [
            customer.get("phoneNumber") or customer.get("normalizedPhone") or "",
            customer.get("secondaryPhoneNumber")
            or customer.get("normalizedSecondaryPhone")
            or "",
        ]
        phone_value = ", ".join(phone for phone in phone_numbers if phone)
        total_minor = order.get("totalAmountMinor", 0) or 0
        delivery_minor = order.get("deliveryFeeMinor", 0) or 0
        paid_minor = order.get("paidAmountMinor", 0) or 0
        sheet.append(
            [
                order.get("waybillNumber", ""),
                order.get("orderNumber") or order.get("id", ""),
                customer.get("name", ""),
                format_address(address),
                address.get("district", ""),
                address.get("city", ""),
                phone_value,
                (total_minor - paid_minor) / 100,
                ", ".join(
                    f"{item.get('name', '')} {item.get('size', '')}".strip()
                    for item in items
                    if item.get("name")
                ),
                (total_minor - delivery_minor) / 100,
            ],
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2, min_col=8, max_col=8):
        row[0].number_format = '#,##0.00'
    for row in sheet.iter_rows(min_row=2, min_col=10, max_col=10):
        row[0].number_format = '#,##0.00'

    for column in sheet.columns:
        maximum_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(
            maximum_length + 2,
            45,
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def export_orders(
    database,
    business_id,
    status=None,
    search=None,
    date_from=None,
    date_to=None,
    courier_id=None,
    order_ids=None,
):
    if not courier_id:
        raise ApiError(
            "courier_required",
            "Choose a courier before exporting orders.",
            422,
        )
    template = get_courier_export_template(database, business_id, courier_id)
    orders = list_orders(
        database, business_id, status=status, search=search,
        date_from=date_from, date_to=date_to, courier_id=courier_id,
    )
    selected_order_ids = {
        str(order_id).strip() for order_id in (order_ids or []) if str(order_id).strip()
    }
    if selected_order_ids:
        orders = [order for order in orders if order.get("id") in selected_order_ids]
    return build_orders_workbook(orders, template_bytes=template.get("content"))


# A fixed document id, so a provider that fails on every message leaves one
# notification rather than hundreds.
AI_STATUS_NOTIFICATION_ID = "ai-status"

AI_FAILURE_NOTIFICATIONS = {
    "configuration": (
        "Chatbot AI is not responding",
        "Customer replies have dropped back to simplified English and are no "
        "longer answering in Sinhala or Tamil. This does not fix itself - the "
        "AI model or API key needs attention.",
    ),
    "rate_limit": (
        "Chatbot AI is rate limited",
        "Some replies are falling back to simplified English while the limit "
        "resets. If this keeps happening the provider plan needs a higher quota.",
    ),
    "unavailable": (
        "Chatbot AI could not be reached",
        "The last request to the AI provider failed. Replies are falling back "
        "to simplified English until it recovers.",
    ),
}

# business_id -> the failure timestamp already written. Avoids a Firestore
# write on every chat message while a provider is down.
_SYNCED_AI_FAILURES = {}


def sync_ai_failure_notification(database, business_id, status):
    """Raise or clear the seller's notification about a broken chatbot AI.

    A provider failure used to reach only the server log, so a seller whose bot
    had quietly dropped to simplified English had no way to find out.
    """
    failure = status.get("failure") if status else None
    marker = failure["at"] if failure else ""

    if _SYNCED_AI_FAILURES.get(business_id) == marker:
        return

    _SYNCED_AI_FAILURES[business_id] = marker
    reference = (
        database.collection("businesses")
        .document(business_id)
        .collection("notifications")
        .document(AI_STATUS_NOTIFICATION_ID)
    )

    if not failure:
        # Recovered. Leaving a stale warning up is how a notification becomes
        # something people learn to ignore.
        reference.delete()
        return

    title, message = AI_FAILURE_NOTIFICATIONS.get(
        failure.get("kind"),
        AI_FAILURE_NOTIFICATIONS["unavailable"],
    )
    reference.set(
        {
            "type": "ai-status",
            "title": title,
            "message": (
                f"{message} Provider: {failure.get('provider', '')}, "
                f"model: {failure.get('model', '')}."
            ),
            "isRead": False,
            "createdAt": firestore.SERVER_TIMESTAMP,
        },
    )


def list_notifications(database, business_id, unread_only=False):
    snapshots = (
        database.collection("businesses")
        .document(business_id)
        .collection("notifications")
        .order_by("createdAt", direction="DESCENDING")
        .limit(50)
        .stream()
    )
    notifications = [serialize_snapshot(snapshot) for snapshot in snapshots]

    if unread_only:
        notifications = [item for item in notifications if not item.get("isRead")]

    return notifications


def mark_notification_read(database, business_id, notification_id):
    reference = (
        database.collection("businesses")
        .document(business_id)
        .collection("notifications")
        .document(notification_id)
    )

    if not reference.get().exists:
        raise ApiError("notification_not_found", "Notification not found.", 404)

    reference.update({"isRead": True, "readAt": firestore.SERVER_TIMESTAMP})
    return serialize_snapshot(reference.get())
